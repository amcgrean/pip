import io
import json
import re
from datetime import date, datetime, timezone

import openpyxl
from sqlalchemy.orm import Session

from app.models.import_job import ImportJob
from app.models.product import Product
from app.models.vendor import Vendor
from app.models.vendor_product_mapping import VendorProductMapping

# Sheets to skip when processing product rows
SKIP_SHEETS = {"summary", "suppliers", "branch coverage"}

# Column header variations that map to species_or_material
BRAND_COLUMN_NAMES = {
    "brand",
    "brand / type",
    "species / material",
    "brand / material",
    "brand / product line",
}

# Optional extra column mappings: spreadsheet header -> Product field
EXTRA_COLUMN_MAP = {
    "color": "finish",
    "grade": "grade",
    "profile / type": "profile",
    "profile": "profile",
    "type": "profile",
    "function": "profile",
}


def _strip_emoji_prefix(sheet_name: str) -> str:
    """Remove leading emoji characters and whitespace from a sheet name."""
    # Emojis are typically in the Unicode ranges above U+2000; strip them plus whitespace
    return re.sub(r"^[^\w]+", "", sheet_name, flags=re.UNICODE).strip()


def _cell_str(value) -> str:
    """Return a stripped string from a cell value, or empty string if None."""
    if value is None:
        return ""
    return str(value).strip()


def _parse_last_sold(value) -> date | None:
    """Parse a last-sold value that may be a datetime, date, or string."""
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value).strip()
    if not text:
        return None
    for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%m/%d/%y", "%d-%b-%Y", "%d-%b-%y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None


def _find_header_row(ws) -> tuple[int, dict[str, int]]:
    """Scan the first 10 rows to locate the header row containing 'SKU'.

    Returns (header_row_number, {lowercase_header: col_index}).
    """
    for row_idx in range(1, min(ws.max_row or 1, 11) + 1):
        for col_idx in range(1, (ws.max_column or 1) + 1):
            val = _cell_str(ws.cell(row=row_idx, column=col_idx).value)
            if val.upper() == "SKU":
                headers: dict[str, int] = {}
                for c in range(1, (ws.max_column or 1) + 1):
                    h = _cell_str(ws.cell(row=row_idx, column=c).value)
                    if h:
                        headers[h.lower()] = c
                return row_idx, headers
    return 0, {}


def _detect_brand_column(headers: dict[str, int]) -> int | None:
    """Find the column index for the brand / species_or_material field."""
    for name in BRAND_COLUMN_NAMES:
        if name in headers:
            return headers[name]
    return None


def _find_vendor_by_name(db: Session, supplier_name: str, vendor_cache: dict[str, Vendor | None]) -> Vendor | None:
    """Look up a Vendor by exact or case-insensitive match on vendor_name, with caching."""
    if not supplier_name:
        return None
    key = supplier_name.lower()
    if key in vendor_cache:
        return vendor_cache[key]

    vendor = db.query(Vendor).filter(Vendor.vendor_name.ilike(supplier_name)).first()
    vendor_cache[key] = vendor
    return vendor


def _find_header_row_by_column(ws, target_header: str) -> tuple[int, dict[str, int]]:
    """Scan first 10 rows for a header row containing the target column name."""
    target = target_header.upper()
    for row_idx in range(1, min(ws.max_row or 1, 11) + 1):
        for col_idx in range(1, (ws.max_column or 1) + 1):
            val = _cell_str(ws.cell(row=row_idx, column=col_idx).value)
            if val.upper() == target:
                headers: dict[str, int] = {}
                for c in range(1, (ws.max_column or 1) + 1):
                    h = _cell_str(ws.cell(row=row_idx, column=c).value)
                    if h:
                        headers[h.lower()] = c
                return row_idx, headers
    return 0, {}


def _process_suppliers_sheet(db: Session, ws) -> dict[str, int]:
    """Process the Suppliers sheet, creating/updating Vendor records.

    Handles shared vendor codes (e.g. multiple LMC suppliers under LMC1000)
    by creating separate vendor records keyed on vendor_name.

    Returns {lowercase_vendor_name: vendor_id} cache.
    """
    header_row, headers = _find_header_row_by_column(ws, "Supplier Code")
    if not header_row:
        return {}

    code_col = headers.get("supplier code")
    name_col = headers.get("supplier name")
    if not code_col or not name_col:
        return {}

    # First pass: collect all (code, name) pairs to detect shared codes
    rows: list[tuple[str, str]] = []
    code_counts: dict[str, int] = {}
    for row_idx in range(header_row + 1, (ws.max_row or header_row) + 1):
        code = _cell_str(ws.cell(row=row_idx, column=code_col).value)
        name = _cell_str(ws.cell(row=row_idx, column=name_col).value)
        if not code or not name:
            continue
        rows.append((code, name))
        code_counts[code] = code_counts.get(code, 0) + 1

    # Second pass: create/update vendors
    # For shared codes (e.g. LMC1000 used by 19 suppliers), each gets a
    # unique vendor_code suffix like LMC1000-01, LMC1000-02 etc.
    vendor_map: dict[str, int] = {}
    code_seq: dict[str, int] = {}  # track sequence for shared codes

    for base_code, name in rows:
        if code_counts[base_code] > 1:
            # Shared code — make unique per supplier name
            seq = code_seq.get(base_code, 0) + 1
            code_seq[base_code] = seq
            unique_code = f"{base_code}-{seq:02d}"
        else:
            unique_code = base_code

        # Look up by name first (in case we already have this supplier)
        vendor = db.query(Vendor).filter(Vendor.vendor_name == name).first()
        if vendor:
            # Update code if needed
            if vendor.vendor_code != unique_code:
                vendor.vendor_code = unique_code
        else:
            # Also check if the unique_code already exists
            vendor = db.query(Vendor).filter(Vendor.vendor_code == unique_code).first()
            if vendor:
                vendor.vendor_name = name
            else:
                vendor = Vendor(vendor_code=unique_code, vendor_name=name, is_active=True)
                db.add(vendor)
                db.flush()

        vendor_map[name.lower()] = vendor.id

    db.commit()
    return vendor_map


def _upsert_product(
    db: Session,
    sku: str,
    values: dict,
) -> tuple[Product, int, int]:
    """Upsert a Product by internal_sku. Returns (product, inserted, updated)."""
    product = db.query(Product).filter(Product.internal_sku == sku).first()
    if product:
        for key, val in values.items():
            setattr(product, key, val)
        return product, 0, 1

    product = Product(internal_sku=sku, **values)
    db.add(product)
    db.flush()
    return product, 1, 0


def _create_vendor_mapping(
    db: Session,
    product: Product,
    vendor: Vendor,
    sku: str,
) -> None:
    """Create a VendorProductMapping if one does not already exist."""
    existing = (
        db.query(VendorProductMapping)
        .filter(
            VendorProductMapping.vendor_id == vendor.id,
            VendorProductMapping.product_id == product.id,
            VendorProductMapping.vendor_sku == sku,
        )
        .first()
    )
    if not existing:
        db.add(
            VendorProductMapping(
                vendor_id=vendor.id,
                product_id=product.id,
                vendor_sku=sku,
            )
        )
        db.flush()


def _process_product_sheet(
    db: Session,
    ws,
    sheet_name: str,
    vendor_cache: dict[str, Vendor | None],
) -> dict:
    """Process a single product category sheet.

    Returns a stats dict with inserted, updated, errored counts and error list.
    """
    category_major = _strip_emoji_prefix(sheet_name)
    header_row, headers = _find_header_row(ws)
    if not header_row or "sku" not in headers:
        return {"category": category_major, "inserted": 0, "updated": 0, "errored": 0, "errors": []}

    sku_col = headers["sku"]
    desc_col = headers.get("description")
    subcat_col = headers.get("subcategory")
    branch_col = headers.get("branches")
    supplier_col = headers.get("supplier")
    last_sold_col = headers.get("last sold")
    brand_col = _detect_brand_column(headers)

    # Detect extra columns
    extra_cols: dict[str, int] = {}
    for header_name, field_name in EXTRA_COLUMN_MAP.items():
        if header_name in headers and field_name not in extra_cols.values():
            extra_cols[header_name] = headers[header_name]

    inserted = 0
    updated = 0
    errored = 0
    errors: list[dict[str, str]] = []

    for row_idx in range(header_row + 1, (ws.max_row or header_row) + 1):
        try:
            raw_sku = _cell_str(ws.cell(row=row_idx, column=sku_col).value)

            # Skip section headers (start with triangle) and blank SKUs
            if not raw_sku or raw_sku.startswith("\u25b8") or raw_sku.startswith(">"):
                continue

            description = _cell_str(ws.cell(row=row_idx, column=desc_col).value) if desc_col else ""
            if not description:
                continue

            product_values: dict = {
                "normalized_name": description,
                "description": description,
                "category_major": category_major,
                "category_minor": _cell_str(ws.cell(row=row_idx, column=subcat_col).value) if subcat_col else None,
                "branch_scope": _cell_str(ws.cell(row=row_idx, column=branch_col).value) if branch_col else None,
                "is_stock_item": True,
                "is_active": True,
                "status": "active",
            }

            # last_sold_date
            if last_sold_col:
                product_values["last_sold_date"] = _parse_last_sold(
                    ws.cell(row=row_idx, column=last_sold_col).value
                )

            # species_or_material from the brand/species column
            if brand_col:
                product_values["species_or_material"] = (
                    _cell_str(ws.cell(row=row_idx, column=brand_col).value) or None
                )

            # Extra columns
            for header_name, col_idx in extra_cols.items():
                field_name = EXTRA_COLUMN_MAP[header_name]
                val = _cell_str(ws.cell(row=row_idx, column=col_idx).value) or None
                # Only set if not already set by a higher-priority column
                if product_values.get(field_name) is None:
                    product_values[field_name] = val

            # Filter out empty-string values to avoid overwriting with ""
            product_values = {k: (v if v != "" else None) for k, v in product_values.items()}
            # Restore required non-None fields
            product_values["normalized_name"] = description
            product_values["is_stock_item"] = True
            product_values["is_active"] = True
            product_values["status"] = "active"

            product, ins, upd = _upsert_product(db, raw_sku, product_values)
            inserted += ins
            updated += upd

            # Vendor mapping
            if supplier_col:
                supplier_name = _cell_str(ws.cell(row=row_idx, column=supplier_col).value)
                if supplier_name:
                    vendor = _find_vendor_by_name(db, supplier_name, vendor_cache)
                    if vendor:
                        _create_vendor_mapping(db, product, vendor, raw_sku)

            # Commit in batches of 200 for performance
            if (inserted + updated) % 200 == 0:
                db.commit()

        except Exception as exc:  # noqa: BLE001
            errored += 1
            errors.append({"row": str(row_idx), "error": str(exc)})
            db.rollback()

    db.commit()
    return {
        "category": category_major,
        "inserted": inserted,
        "updated": updated,
        "errored": errored,
        "errors": errors,
    }


def process_stock_catalog(db: Session, file_content: bytes, file_name: str, user_id: int) -> ImportJob:
    """Parse a Beisser Stock Catalog XLSX and upsert products, vendors, and mappings.

    Returns the completed ImportJob record.
    """
    job = ImportJob(
        source_type="stock_catalog",
        file_name=file_name,
        status="processing",
        created_by=user_id,
    )
    db.add(job)
    db.commit()
    db.refresh(job)

    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_content), read_only=True, data_only=True)
    except Exception as exc:
        job.status = "failed"
        job.error_log = f"Failed to open workbook: {exc}"
        db.commit()
        return job

    # 1. Process Suppliers sheet first
    vendor_cache: dict[str, Vendor | None] = {}
    for ws_name in wb.sheetnames:
        if "supplier" in ws_name.lower():
            vendor_id_map = _process_suppliers_sheet(db, wb[ws_name])
            # Pre-populate the vendor_cache from the supplier map
            for name_lower, vid in vendor_id_map.items():
                vendor_cache[name_lower] = db.query(Vendor).filter(Vendor.id == vid).first()
            break

    # 2. Process each product category sheet
    total_inserted = 0
    total_updated = 0
    total_errored = 0
    total_rows = 0
    all_errors: list[dict[str, str]] = []
    category_stats: list[dict] = []

    for ws_name in wb.sheetnames:
        clean_name = _strip_emoji_prefix(ws_name).lower().strip()
        if clean_name in SKIP_SHEETS:
            continue

        ws = wb[ws_name]
        stats = _process_product_sheet(db, ws, ws_name, vendor_cache)
        sheet_total = stats["inserted"] + stats["updated"] + stats["errored"]
        if sheet_total == 0:
            continue

        total_inserted += stats["inserted"]
        total_updated += stats["updated"]
        total_errored += stats["errored"]
        total_rows += sheet_total
        all_errors.extend(
            {**e, "sheet": stats["category"]} for e in stats["errors"]
        )
        category_stats.append(
            {
                "category": stats["category"],
                "inserted": stats["inserted"],
                "updated": stats["updated"],
                "errored": stats["errored"],
            }
        )

    wb.close()

    # 3. Finalize ImportJob
    job.total_rows = total_rows
    job.inserted_rows = total_inserted
    job.updated_rows = total_updated
    job.error_rows = total_errored
    job.error_log = json.dumps({"errors": all_errors[:100], "by_category": category_stats}) if all_errors else json.dumps({"by_category": category_stats})
    job.status = "completed_with_errors" if total_errored else "completed"
    job.completed_at = datetime.now(timezone.utc)
    db.commit()
    db.refresh(job)
    return job
